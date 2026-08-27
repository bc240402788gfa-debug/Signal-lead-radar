"""
Blazeup Signal - Telegram bot

Send /scan <location> [| keyword1, keyword2, ...] and get back a ranked lead list
as a downloadable .xlsx file, right in the chat.

--- Setup ---
1. pip install python-telegram-bot httpx openpyxl
2. Get a bot token from @BotFather on Telegram (message it, /newbot, follow the prompts).
3. Set the token as an environment variable. NEVER hardcode it in this file or commit
   it to GitHub -- that repo is public.
     Windows (PowerShell, run once, then open a new terminal):
       setx BLAZEUP_BOT_TOKEN "your-token-here"
4. Run: python telegram_bot.py
5. For 24/7 operation, add it to Task Scheduler exactly like the web server and the
   trading bots: Action = python.exe, Arguments = the full path to this file,
   Trigger = At startup, "run whether user is logged on or not". Long polling means
   this needs no inbound port opened, unlike the web server.

sectors.py must be in the same folder -- this script imports it directly so the
category list and base scores are driven by the same source of truth as the website.

Known limitation worth knowing: the geocode/Overpass/scoring logic below is a SECOND,
independent implementation of what app.js does in the browser. A Telegram bot has no
browser to run the JS in, so this had to be ported to Python by hand. If the scoring
model in app.js changes again, this file needs the same change made separately --
they will not stay in sync automatically.
"""

import os
import re
import logging
from io import BytesIO

import httpx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes

import sectors  # sectors.py, same folder

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("blazeup-bot")

RADIUS_KM_DEFAULT = 8
OVERPASS_TIMEOUT_SECONDS = 40

SECTOR_BY_ID = {s["id"]: s for s in sectors.SECTORS}


def resolve_sectors(keyword_text):
    """User-typed keywords -> a set of active sector ids. Falls back to the same
    default-on set the website uses if no keywords are given or none match."""
    default_ids = {s["id"] for s in sectors.SECTORS if s["defaultOn"]}
    if not keyword_text or not keyword_text.strip():
        return default_ids
    keywords = [k.strip().lower() for k in keyword_text.split(",") if k.strip()]
    active = set()
    for kw in keywords:
        for s in sectors.SECTORS:
            if kw in s["label"].lower() or kw in s["id"].lower():
                active.add(s["id"])
    return active or default_ids


def build_overpass_query(lat, lon, radius_m, active_ids):
    clauses = []
    for row in sectors.TAGMAP:
        key, val, sector_id = row[0], row[1], row[2]
        brand_req = row[3] if len(row) > 3 else False
        if sector_id not in active_ids:
            continue
        brand_filter = '["brand"]' if brand_req else ""
        clauses.append('  node["%s"="%s"]%s(around:%d,%s,%s);' % (key, val, brand_filter, radius_m, lat, lon))
        clauses.append('  way["%s"="%s"]%s(around:%d,%s,%s);' % (key, val, brand_filter, radius_m, lat, lon))
    body = "\n".join(clauses)
    return "[out:json][timeout:%d];\n(\n%s\n);\nout center 900;" % (OVERPASS_TIMEOUT_SECONDS, body)


async def geocode(client, location):
    resp = await client.get(
        "https://nominatim.openstreetmap.org/search",
        params={"format": "json", "limit": 1, "q": location},
    )
    resp.raise_for_status()
    data = resp.json()
    if not data:
        return None
    return {"lat": float(data[0]["lat"]), "lon": float(data[0]["lon"]), "label": data[0]["display_name"]}


async def fetch_overpass(client, query):
    resp = await client.post("https://overpass-api.de/api/interpreter", data={"data": query})
    resp.raise_for_status()
    data = resp.json()
    if data.get("remark"):
        raise RuntimeError("QUERY_TOO_LARGE: " + data["remark"])
    return data.get("elements", [])


def process_elements(elements):
    """Mirrors app.js's processElements/scoring exactly -- see the module docstring
    about this being a hand-kept-in-sync duplicate, not a shared source."""
    seen = {}
    for el in elements:
        tags = el.get("tags") or {}
        cat = None
        for row in sectors.TAGMAP:
            key, val, sector_id = row[0], row[1], row[2]
            if tags.get(key) == val:
                cat = SECTOR_BY_ID.get(sector_id)
                break
        if not cat:
            continue
        name = tags.get("name") or tags.get("name:en") or tags.get("brand")
        if not name:
            continue
        key_id = "%s/%s" % (el.get("type"), el.get("id"))
        if key_id in seen:
            continue
        if el.get("type") == "node":
            lat, lon = el.get("lat"), el.get("lon")
        else:
            center = el.get("center") or {}
            lat, lon = center.get("lat"), center.get("lon")
        if lat is None or lon is None:
            continue

        website = tags.get("website") or tags.get("contact:website")
        phone = tags.get("phone") or tags.get("contact:phone")
        facebook = tags.get("contact:facebook") or tags.get("facebook")
        instagram = tags.get("contact:instagram") or tags.get("instagram")
        addr_parts = [tags.get("addr:housenumber"), tags.get("addr:street"), tags.get("addr:city")]
        address = " ".join(p for p in addr_parts if p) or None

        has_social = bool(facebook or instagram)
        social_count = (1 if facebook else 0) + (1 if instagram else 0)
        score = cat["base"]
        if tags.get("brand"):
            score += 5
        if has_social:
            score += 14
        if social_count == 2:
            score += 4
        if phone:
            score += 5
        if tags.get("addr:street"):
            score += 3
        if website and not has_social:
            score += 4
        score = max(0, min(100, score))

        seen[key_id] = {
            "name": name, "category": cat["label"], "why": cat["why"],
            "website": website, "phone": phone, "facebook": facebook, "instagram": instagram,
            "address": address, "score": score, "has_social": has_social,
        }
    leads = list(seen.values())
    leads.sort(key=lambda l: (-l["score"], not l["has_social"], l["name"]))
    return leads


def build_xlsx(leads):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(["Rank", "Name", "Category", "Score", "Social Presence", "Address", "Phone", "Website", "Facebook", "Instagram", "Why"])
    for i, l in enumerate(leads, start=1):
        ws.append([
            i, l["name"], l["category"], l["score"], "Yes" if l["has_social"] else "No",
            l["address"] or "", l["phone"] or "", l["website"] or "", l["facebook"] or "", l["instagram"] or "", l["why"],
        ])
    for idx, w in enumerate([5, 24, 22, 7, 14, 30, 16, 26, 22, 22, 60], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def slug(s):
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s[:40] or "leads"


async def scan_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = " ".join(context.args) if context.args else ""
    if not text.strip():
        await update.message.reply_text(
            "Usage:\n"
            "/scan <location>\n"
            "/scan <location> | keyword1, keyword2\n\n"
            "Example: /scan Lahore, Pakistan\n"
            "Example: /scan Lahore, Pakistan | restaurants, beauty\n\n"
            "Send /sectors to see available keywords."
        )
        return

    if "|" in text:
        location, keyword_text = text.split("|", 1)
    else:
        location, keyword_text = text, ""
    location = location.strip()
    active_ids = resolve_sectors(keyword_text)

    await update.message.reply_text("Scanning %s (%d categories)..." % (location, len(active_ids)))

    async with httpx.AsyncClient(timeout=OVERPASS_TIMEOUT_SECONDS + 15) as client:
        try:
            geo = await geocode(client, location)
        except Exception:
            log.exception("geocode failed")
            await update.message.reply_text("Could not reach the location service. Try again shortly.")
            return
        if not geo:
            await update.message.reply_text("Location not found -- try a nearby city or add a country name.")
            return

        query = build_overpass_query(geo["lat"], geo["lon"], RADIUS_KM_DEFAULT * 1000, active_ids)
        try:
            elements = await fetch_overpass(client, query)
        except RuntimeError as e:
            if str(e).startswith("QUERY_TOO_LARGE"):
                await update.message.reply_text("Query too large for that area/category mix -- try fewer keywords.")
            else:
                await update.message.reply_text("Overpass request failed. Try again shortly.")
            return
        except Exception:
            log.exception("overpass fetch failed")
            await update.message.reply_text("Could not reach the map data service. Try again shortly.")
            return

    leads = process_elements(elements)
    if not leads:
        await update.message.reply_text("No matches within %dkm of %s." % (RADIUS_KM_DEFAULT, geo["label"]))
        return

    xlsx_buf = build_xlsx(leads)
    top5 = "\n".join("%d. %s (%d) - %s" % (i + 1, l["name"], l["score"], l["category"]) for i, l in enumerate(leads[:5]))
    await update.message.reply_document(
        document=xlsx_buf,
        filename="blazeup-leads-%s.xlsx" % slug(geo["label"]),
        caption="%d leads within %dkm of %s.\n\nTop 5:\n%s" % (len(leads), RADIUS_KM_DEFAULT, geo["label"], top5),
    )


async def sectors_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lines = ["- %s%s" % (s["label"], " (default)" if s["defaultOn"] else "") for s in sectors.SECTORS]
    await update.message.reply_text("Available categories:\n" + "\n".join(lines))


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Blazeup Signal bot.\n\n"
        "/scan <location> - find leads using default categories\n"
        "/scan <location> | keyword1, keyword2 - filter to specific categories\n"
        "/sectors - list available categories\n\n"
        "Radius is fixed at %dkm for now." % RADIUS_KM_DEFAULT
    )


def main():
    token = os.environ.get("BLAZEUP_BOT_TOKEN")
    if not token:
        raise SystemExit(
            "BLAZEUP_BOT_TOKEN environment variable not set. Get a token from @BotFather "
            "on Telegram, then set it as an env var -- never hardcode it in this file or "
            "commit it to GitHub."
        )
    app = Application.builder().token(token).build()
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("scan", scan_command))
    app.add_handler(CommandHandler("sectors", sectors_command))
    log.info("Blazeup Signal bot starting (long polling)...")
    app.run_polling()


if __name__ == "__main__":
    main()
